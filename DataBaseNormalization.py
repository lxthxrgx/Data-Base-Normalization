import sqlite3 as sq
import re
from openpyxl import *
import io

sqlite_connection = sq.connect('ltx.db')
cursor = sqlite_connection.cursor() 

workbook = load_workbook(r'excel\excel sources\montransactions.xlsx')
sheet = workbook['123']
cursor.execute('DROP TABLE IF EXISTS mon')
cursor.execute('''CREATE TEMPORARY TABLE mon (
    Id INTEGER PRIMARY KEY,
	Cudnum	TEXT,
	Koatyy	TEXT,
	Area	TEXT,
	District	TEXT,
	Settlement	TEXT,
	City	TEXT,
	Street	TEXT,
	TGName	TEXT,
	PurposeOfTheAssignment	TEXT,
    PurposeOfTheAssignmentNormal TEXT,
	NameOfTheSite	TEXT,
	TransactionType	TEXT,
	Price	TEXT,
	OwnershipType	TEXT,
	RegistrationDate	TEXT,
	RegistrationNumber	TEXT,
	ValueNGO	TEXT,
	EvaluationDate	TEXT
)''')

last_row = sheet.max_row
while sheet.cell(row=last_row, column=2).value is None and last_row > 1:
    last_row -= 1
arc = last_row+1


for i in range(2, 470864):
     a = sheet.cell(row=i, column = 2).value
     b = sheet.cell(row=i, column = 3).value
     c = sheet.cell(row=i, column = 4).value
     d = sheet.cell(row=i, column = 5).value
     e = sheet.cell(row=i, column = 6).value
     f = sheet.cell(row=i, column = 7).value
     g = sheet.cell(row=i, column = 8).value
     h = sheet.cell(row=i, column = 9).value
     x = sheet.cell(row=i, column = 10).value
     j = sheet.cell(row=i, column = 11).value
     k = sheet.cell(row=i, column = 12).value
     l = sheet.cell(row=i, column = 13).value
     m = sheet.cell(row=i, column = 14).value
     n = sheet.cell(row=i, column = 15).value
     o = sheet.cell(row=i, column = 16).value
     p = sheet.cell(row=i, column = 17).value
     q = sheet.cell(row=i, column = 18).value
     cursor.execute('''INSERT INTO mon (
        Cudnum,
        Koatyy,
        Area,
        District,
        Settlement,
        City,
        Street,
        TGName,
        PurposeOfTheAssignment,
        NameOfTheSite,
        TransactionType,
        Price,
        OwnershipType,
        RegistrationDate,
        RegistrationNumber,
        ValueNGO,
        EvaluationDate) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',(a,b,c,d,e,f,g,h,x,j,k,l,m,n,o,p,q,))

sqlite_connection.commit()
workbook.close()

cursor.execute('CREATE TEMPORARY TABLE "dictionary" (fcdata	NUMERIC, nmdata	NUMERIC)')
cursor.execute('INSERT INTO dictionary(fcdata, nmdata) SELECT PurposeOfTheAssignment, NULL FROM mon')

reader = io.open(r'C:\Users\3d\Desktop\innertext1.txt', 'r', encoding='utf-8')
text = reader.read()

filter_file = r'[+]?\d{2}\.\d{2}'

normalized_data_from_file_list = re.findall(filter_file, text)

normalized_data_from_file_list = list(dict.fromkeys(normalized_data_from_file_list))
result = ''.join(normalized_data_from_file_list)

float_normalized_data_from_file_list = []
for number in normalized_data_from_file_list:
    float_number = float(number)
    if float_number <= 19:
        float_normalized_data_from_file_list.append(float_number)

for i, a in enumerate(float_normalized_data_from_file_list[1:], start=2):
    cursor.execute(f"UPDATE dictionary SET nmdata = ? WHERE ROWID = {i};", (a,))

cursor.execute("""CREATE TEMPORARY TABLE mon2(
    Id INTEGER PRIMARY KEY,
	Cudnum TEXT,
	Koatyy TEXT,
	Area TEXT,
	District TEXT,
	Settlement TEXT,
	City TEXT,
	Street TEXT,
	TGName TEXT,
	PurposeOfTheAssignment REAL,
    PurposeOfTheAssignmentNormal REAL,
	NameOfTheSite TEXT,
	TransactionType TEXT,
	Price TEXT,
	OwnershipType TEXT,
	RegistrationDate TEXT,
	RegistrationNumber TEXT,
	ValueNGO TEXT,
	EvaluationDate TEXT
)""")

cursor.execute("""INSERT INTO mon2(
    Id ,
    Cudnum,
    Koatyy,
    Area,
    District,
    Settlement,
    City,Street,
    TGName,
    PurposeOfTheAssignment,
    PurposeOfTheAssignmentNormal,
    NameOfTheSite,
    TransactionType,
    Price,
    OwnershipType,
    RegistrationDate,
    RegistrationNumber,
    ValueNGO,
    EvaluationDate) SELECT 
    Id ,
    Cudnum,
    Koatyy,
    Area,
    District,
    Settlement,
    City,Street,
    TGName,
    PurposeOfTheAssignment,
    PurposeOfTheAssignmentNormal,
    NameOfTheSite,
    TransactionType,
    Price,
    OwnershipType,
    RegistrationDate,
    RegistrationNumber,
    ValueNGO,
    EvaluationDate
    FROM mon""")

cursor.execute('CREATE TABLE IF NOT EXISTS monitoring AS SELECT * FROM mon2')

cursor.execute('''UPDATE monitoring SET PurposeOfTheAssignmentNormal = (SELECT nmdata FROM dictionary WHERE monitoring.PurposeOfTheAssignment LIKE '%' || dictionary.nmdata || '%') WHERE PurposeOfTheAssignmentNormal IS NULL''')

reader.close()

cursor.execute('SELECT PurposeOfTheAssignmentNormal,COUNT(PurposeOfTheAssignmentNormal), Area, Price, ValueNGO, AVG(Price/Area) AS Pa, AVG(Price/ValueNGO) AS Pv FROM monitoring GROUP BY PurposeOfTheAssignmentNormal')
rows = cursor.fetchall()


workbook = load_workbook('F:\Prog\Py\excel\excel draft\ltx.xlsx')
sheet = workbook['123']

row_num = 8
for row in rows:
    purpose = row[0]
    count = row[1]
    price = row[3]
    pa = row[5]
    pv = row[6]

    sheet.cell(row=row_num, column=2).value = purpose
    sheet.cell(row=row_num, column=3).value = count
    sheet.cell(row=row_num, column=6).value = price
    sheet.cell(row=row_num, column=5).value = pa
    sheet.cell(row=row_num, column=7).value = pv

    row_num += 1


workbook.save('F:\Prog\Py\excel\excel draft\ltx.xlsx')


cursor.close()

sqlite_connection.commit()
cursor.close()
